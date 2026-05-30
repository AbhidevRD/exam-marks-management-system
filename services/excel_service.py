from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

def generate_all_students_excel(students_data, subject_names):

    wb = Workbook()
    ws = wb.active
    ws.title = "All Students Marksheet"

    # Header
    headers = ["Sr No", "Name"] + subject_names + ["Total", "Average", "Grade", "Result"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Insert student rows
    for index, student in enumerate(students_data, start=1):

        subject_values = [student["subjects"][sub] for sub in subject_names]

        row = [
            index,
            student["name"],
            *subject_values,
            student["total"],
            student["average"],
            student["grade"],
            student["result"]
        ]

        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_subject_wise_excel(students_data, subject_name):
    """Generate Excel with marks for all students in a specific subject"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject_name} Marksheet"

    # Header
    headers = ["Sr No", "Student Name", f"{subject_name} Marks"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Insert student rows
    for index, student in enumerate(students_data, start=1):
        marks = student["subjects"].get(subject_name, 0)
        
        row = [
            index,
            student["name"],
            marks
        ]
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_exam_wise_excel(students_data, subject_names, exam_name):
    """Generate Excel with all marks for all students in a specific exam"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{exam_name}"

    # Header
    headers = ["Sr No", "Name"] + subject_names + ["Total", "Average", "Grade", "Result"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Insert student rows
    for index, student in enumerate(students_data, start=1):

        subject_values = [student["subjects"][sub] for sub in subject_names]

        row = [
            index,
            student["name"],
            *subject_values,
            student["total"],
            student["average"],
            student["grade"],
            student["result"]
        ]

        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_student_wise_excel(student_data, subject_names):
    """Generate Excel for a single student with all marks"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Marksheet"

    # Title
    ws.append([f"Student Name: {student_data['name']}"])
    ws.append([])

    # Headers
    headers = ["Subject", "Marks"]
    ws.append(headers)

    # Style header
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Insert subject rows
    for subject in subject_names:
        marks = student_data["subjects"].get(subject, 0)
        row = [subject, marks]
        ws.append(row)

    # Summary section
    ws.append([])
    ws.append(["Total", student_data["total"]])
    ws.append(["Average", student_data["average"]])
    ws.append(["Grade", student_data["grade"]])
    ws.append(["Result", student_data["result"]])

    # Style summary
    for row in ws.iter_rows(min_row=ws.max_row-3, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
